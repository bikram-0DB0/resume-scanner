import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

def create_output_directory(base_name="resume_extraction_results"):
    """Create a timestamped output directory"""
    output_dir = f"{base_name}"
    
    os.makedirs(output_dir, exist_ok=True)
    return output_dir

def save_to_csv(data, output_file='extracted_resume_data.csv'):
    """Save extracted data to CSV file with enhanced formatting"""
    if not data:
        print("⚠️ No data to save to CSV")
        return
    
    try:
        df = pd.DataFrame(data)
        
        desired_order = ['Name', 'Email', 'Phone', 'Skills', 
                        'Work Experience', 'Education', 'Projects', 'Hobbies', 'Qualities']
        
        existing_columns = [col for col in desired_order if col in df.columns]
        remaining_columns = [col for col in df.columns if col not in existing_columns and col != 'Resume Name']
        
        column_order = existing_columns + remaining_columns
        if 'Resume Name' in df.columns:
            column_order.append('Resume Name')
        
        df = df[column_order]
        
        df.to_csv(output_file, index=False, encoding='utf-8')
        print(f"✅ CSV saved successfully: {output_file}")
        
    except Exception as e:
        print(f"❌ Error saving CSV: {str(e)}")

def save_to_excel(data, output_file='extracted_resume_data.xlsx'):
    """Save extracted data to Excel file with enhanced formatting"""
    if not data:
        print("⚠️ No data to save to Excel")
        return
    
    try:
        df = pd.DataFrame(data)
        
        desired_order = ['Name', 'Email', 'Phone', 'Skills', 
                        'Work Experience', 'Education', 'Projects', 'Hobbies', 'Qualities']
        
        existing_columns = [col for col in desired_order if col in df.columns]
        remaining_columns = [col for col in df.columns if col not in existing_columns and col != 'Resume Name']
        
        column_order = existing_columns + remaining_columns
        if 'Resume Name' in df.columns:
            column_order.append('Resume Name')
        
        df = df[column_order]
        
        df.to_excel(output_file, index=False, engine='openpyxl')
        
        format_excel_file(output_file)
        print(f"✅ Excel saved successfully: {output_file}")
        
    except Exception as e:
        print(f"❌ Error saving Excel: {str(e)}")

def format_excel_file(output_file):
    """Apply enhanced formatting to Excel file"""
    try:
        wb = load_workbook(output_file)
        ws = wb.active
        
        header_font = Font(bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column
            column_letter = get_column_letter(column)
            
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, 12), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        data_alignment = Alignment(vertical='top', wrap_text=True)
        
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = data_alignment
                
                ws.row_dimensions[cell.row].height = 60
        
        ws.freeze_panes = 'A2'
        
        ws.auto_filter.ref = ws.dimensions
        
        wb.save(output_file)
        
    except Exception as e:
        print(f"⚠️ Warning: Could not apply advanced formatting: {str(e)}")

def create_summary_report(data, output_dir):
    """Create a summary report of the extraction process"""
    if not data:
        return
    
    try:
        summary_file = os.path.join(output_dir, "extraction_summary.txt")
        
        with open(summary_file, 'w', encoding='utf-8') as f:
            f.write("RESUME EXTRACTION SUMMARY REPORT\n")
            f.write("=" * 50 + "\n\n")
            f.write(f"Total resumes processed: {len(data)}\n\n")
            
            f.write("FIELD COMPLETION STATISTICS:\n")
            f.write("-" * 30 + "\n")
            
            fields = ['Name', 'Email', 'Phone', 'Skills', 'Work Experience', 
                     'Education', 'Projects', 'Hobbies', 'Qualities']
            
            for field in fields:
                completed = sum(1 for item in data if item.get(field))
                percentage = (completed / len(data)) * 100
                f.write(f"{field}: {completed}/{len(data)} ({percentage:.1f}%)\n")
            
            f.write(f"\nSummary saved to: {summary_file}")
        
        print(f"📊 Summary report saved: {summary_file}")
        
    except Exception as e:
        print(f"⚠️ Could not create summary report: {str(e)}")

def validate_output_directory(output_dir):
    """Validate and create output directory if needed"""
    try:
        if not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
            print(f"📁 Created output directory: {output_dir}")
        
        test_file = os.path.join(output_dir, "test_write.tmp")
        with open(test_file, 'w') as f:
            f.write("test")
        os.remove(test_file)
        
        return True
        
    except Exception as e:
        print(f"❌ Cannot write to output directory: {str(e)}")
        return False
